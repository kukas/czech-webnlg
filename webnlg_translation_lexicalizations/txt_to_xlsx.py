from argparse import ArgumentParser
from openpyxl import Workbook, load_workbook


def main():
    ap = ArgumentParser(
        description="Convert txt files into xlsx and back. One line converts to one row. "
        "This is different from converting from csv because csv must be parsed and may contain multiple lines per one cell value"
    )
    ap.add_argument(
        "input_file",
        type=str,
        help="Input file (.txt or .xlsx)",
    )
    ap.add_argument(
        "output_file",
        type=str,
        help="Output file (.txt or .xlsx)",
    )
    args = ap.parse_args()

    input_extension = args.input_file.split(".")[-1]
    output_extension = args.output_file.split(".")[-1]

    if input_extension not in ["txt", "xlsx"]:
        raise ValueError("Invalid input file extension")
    if output_extension not in ["txt", "xlsx"]:
        raise ValueError("Invalid output file extension")
    if input_extension == output_extension:
        raise ValueError("Input and output extensions must be different")

    if input_extension == "txt":
        wb = Workbook()

        # grab the active worksheet
        ws = wb.active

        with open(args.input_file, encoding="utf-8") as f:
            lines = [line.strip() for line in f.readlines()]

            for line in lines:
                ws.append([line])

        # Save the file
        wb.save(args.output_file)

    elif input_extension == "xlsx":
        wb = load_workbook(args.input_file)
        ws = wb.active
        with open(args.output_file, "w", encoding="utf-8") as f:
            for v in ws.values:
                assert v[0] is not None, f"Empty cell found {ws}"
                f.write(v[0].strip() + "\n")
                    
            # f.writelines(v[0].strip() + "\n" for v in ws.values)


if __name__ == "__main__":
    main()
